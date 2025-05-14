import streamlit as st
import pandas as pd
import datetime
from google_sheets import connect_to_sheet, read_sheet_as_df, write_df_to_sheet
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# === Banner ===
st.image("banner_makaboom.png", use_container_width=True)

# === Validación de acceso ===
def validar_clave():
    if st.session_state.pin_clave == st.secrets["security"]["pin"]:
        st.session_state.acceso_autorizado = True
    else:
        st.error("PIN incorrecto.")


if "acceso_autorizado" not in st.session_state:
    st.session_state.acceso_autorizado = False

if not st.session_state.acceso_autorizado:
    st.text_input("Ingresa tu PIN:", type="password", key="pin_clave", on_change=validar_clave)
    st.stop()

# === Conexión con Google Sheets ===
SHEET_KEY = "1OPCAwKXoEHBmagpvkhntywqkAit7178pZv3ptXd9d9w"
sheet = connect_to_sheet(st.secrets["credentials"], SHEET_KEY)

#Función obtener mes  y año siguiente
def obtener_mes_siguiente(mes_actual, año_actual):
    if mes_actual == 12:
        return 1, año_actual + 1
    else:
        return mes_actual + 1, año_actual
    

# === Selección de mes y año ===
today = datetime.date.today()
col1, col2, col3 = st.columns(3)
with col1:
    mes = st.selectbox("Mes", list(range(1, 13)), index=today.month - 1, key="mes_selector")
with col2:
    año = st.selectbox("Año", list(range(2024, 2031)), index=1, key="año_selector")
with col3:
    st.markdown("<br>", unsafe_allow_html=True)  # Esto empuja el botón hacia abajo
    if st.button("➡️ Ir a nuevo mes", help="Duplicar datos al mes siguiente"):
        nuevo_mes, nuevo_año = obtener_mes_siguiente(mes, año)
        st.toast(f"Creando datos para {nuevo_mes}/{nuevo_año}...")

        hojas = {
            "Ingresos": {},
            "Provisiones": {"se_usó": "No", "monto_usado": 0},
            "Gastos Fijos": {"estado": "pendiente"},
            "Ahorros": {},
            "Reservas Familiares": {}
        }

        for hoja, ajustes in hojas.items():
            try:
                df = read_sheet_as_df(sheet, hoja)
                df_origen = df[(df["mes"] == mes) & (df["año"] == año)].copy()
                if df_origen.empty:
                    st.toast(f"No hay datos en {hoja} para copiar.")
                    continue

                df_origen["mes"] = nuevo_mes
                df_origen["año"] = nuevo_año

                for col, val in ajustes.items():
                    if col in df_origen.columns:
                        df_origen[col] = val

                df_sin_nuevo = df[~((df["mes"] == nuevo_mes) & (df["año"] == nuevo_año))]
                df_final = pd.concat([df_sin_nuevo, df_origen], ignore_index=True)
                write_df_to_sheet(sheet, hoja, df_final)
                st.toast(f"{hoja} copiado correctamente.")
            except Exception as e:
                st.toast(f"Error al copiar {hoja}: {e}")
        st.session_state["mes_selector"] = nuevo_mes
        st.session_state["año_selector"] = nuevo_año

# === Lectura centralizada de hojas ===
hojas = ["Ingresos", "Gastos Fijos", "Deudas", "Provisiones", "Ahorros", "Reservas Familiares"]
df_hojas = {}
for hoja in hojas:
    try:
        df_hojas[hoja] = read_sheet_as_df(sheet, hoja)
    except:
        df_hojas[hoja] = pd.DataFrame()

# === Leer cuentas ===
try:
    df_cuentas = read_sheet_as_df(sheet, "Cuentas")
    lista_cuentas = df_cuentas["nombre_cuenta"].dropna().unique().tolist()
except:
    lista_cuentas = []

# === Función para mostrar y guardar editor ===
def mostrar_editor(nombre_hoja, columnas_dropdown=None):
    try:
        df = read_sheet_as_df(sheet, nombre_hoja)
    except:
        st.warning(f"No se pudo cargar la hoja '{nombre_hoja}'")
        return

    tiene_mes_anio = "mes" in df.columns and "año" in df.columns
    df_filtrado = df[(df["mes"] == mes) & (df["año"] == año)] if tiene_mes_anio else df.copy()
    st.subheader(f"{nombre_hoja} ({mes}/{año})" if tiene_mes_anio else nombre_hoja)

    columnas_ocultas = ["mes", "año"]
    columnas_visibles = [c for c in df_filtrado.columns if c not in columnas_ocultas]

    edited_df = st.data_editor(
        df_filtrado[columnas_visibles],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            col: st.column_config.SelectboxColumn("Cuenta", options=lista_cuentas, required=True)
            for col in (columnas_dropdown or []) if col in columnas_visibles
        }
    )

    if st.button(f"💾 Guardar cambios en {nombre_hoja}", key=f"save_{nombre_hoja}"):
        if tiene_mes_anio:
            edited_df["mes"] = mes
            edited_df["año"] = año
            df_sin_filtro = df[~((df["mes"] == mes) & (df["año"] == año))]
            df_final = pd.concat([df_sin_filtro, edited_df], ignore_index=True)
        else:
            df_final = edited_df
        write_df_to_sheet(sheet, nombre_hoja, df_final)
        st.success(f"{nombre_hoja} actualizado correctamente.")

# === Tabs para mostrar categorías dentro de "Datos Detallados" ===
main_tabs = st.tabs(["📊 Resumen General",
                    "🔔 Alertas",
                    "📋 Datos Detallados",
                    "📈 Reportes y Análisis",
                    "🧮 Simulador"])


with main_tabs[0]:
    st.subheader("📊 Resumen General")
    #=== RESUMEN GRAL
    try:
        # === Leer hojas ===
        df_ing = read_sheet_as_df(sheet, "Ingresos")
        df_gastos = read_sheet_as_df(sheet, "Gastos Fijos")
        df_deudas = read_sheet_as_df(sheet, "Deudas")
        df_provisiones = read_sheet_as_df(sheet, "Provisiones")
        df_ahorros = read_sheet_as_df(sheet, "Ahorros")

        # === Filtros por mes y año ===
        ing_mes = df_ing[(df_ing["mes"] == mes) & (df_ing["año"] == año)]
        gf_pagados = df_gastos[(df_gastos["mes"] == mes) & (df_gastos["año"] == año) & (df_gastos["estado"].str.lower() == "pagado")]
        deudas_mes = df_deudas[(df_deudas["mes"] == mes) & (df_deudas["año"] == año)]
        provisiones_mes = df_provisiones[(df_provisiones["mes"] == mes) & (df_provisiones["año"] == año)]
        ahorros_mes = df_ahorros[(df_ahorros["mes"] == mes) & (df_ahorros["año"] == año)]

        # === Cálculos ===
        total_ingresos = ing_mes["monto"].sum()
        gasto_deudas = (deudas_mes["monto_cuota"] * deudas_mes["cuotas_mes"]).sum()
        gasto_normal = gf_pagados["monto"].sum() + gasto_deudas
        gasto_provisiones = provisiones_mes["monto_usado"].sum()
        gasto_ahorros = ahorros_mes["monto_retirado"].sum()
        gasto_total = gasto_normal + gasto_provisiones + gasto_ahorros

        provisiones_guardadas = provisiones_mes["monto"].sum()
        ahorros_guardados = ahorros_mes["monto_ingreso"].sum()

        saldo_real = total_ingresos - gasto_total - provisiones_guardadas - ahorros_guardados

        # === Mostrar métricas ===
        col1, col2, col3 = st.columns(3)
        col1.metric("💰 Ingresos Totales", f"${total_ingresos:,.0f}")
        col2.metric("💸 Gasto Total (todos los orígenes)", f"${gasto_total:,.0f}")
        col3.metric("🧮 Saldo Disponible Real", f"${saldo_real:,.0f}")

        # === Barra de progreso de uso del ingreso mensual ===
        if total_ingresos > 0:
            porcentaje_gasto = min(gasto_total / total_ingresos, 1.0)
            st.progress(porcentaje_gasto, text=f"{porcentaje_gasto * 100:.1f}% del ingreso mensual gastado")
        else:
            st.info("Aún no se han registrado ingresos para este mes.")

        # === Distribución del gasto mensual por origen ===
        st.markdown("# 💸 Distribución del gasto mensual por origen")
        col1, col2, col3 = st.columns(3)
        col1.metric("🧾 Desde Ingresos Normales", f"${gasto_normal:,.0f}")
        col2.metric("🏷️ Desde Provisiones", f"${gasto_provisiones:,.0f}")
        col3.metric("🏦 Desde Ahorros", f"${gasto_ahorros:,.0f}")
        st.caption(f"💼 Gasto total del mes (sumado): ${gasto_total:,.0f}")


        # === Barra de progreso de uso del ingreso mensual ===
        if total_ingresos > 0:
            porcentaje_gasto = min(gasto_total / total_ingresos, 1.0)
            st.progress(porcentaje_gasto, text=f"{porcentaje_gasto * 100:.1f}% del ingreso mensual gastado")
        else:
            st.info("Aún no se han registrado ingresos para este mes.")

    except Exception as e:
        st.warning("No se pudo calcular el resumen financiero.")
        st.text(f"Error: {e}")
    
with main_tabs[1]:
    st.subheader("🔔 Alertas")

    try:
        df_ing = read_sheet_as_df(sheet, "Ingresos")
        df_gastos = read_sheet_as_df(sheet, "Gastos Fijos")
        df_deudas = read_sheet_as_df(sheet, "Deudas")
        df_provisiones = read_sheet_as_df(sheet, "Provisiones")
        df_ahorros = read_sheet_as_df(sheet, "Ahorros")

        # === Filtrar por mes y año
        ing_mes = df_ing[(df_ing["mes"] == mes) & (df_ing["año"] == año)]
        gastos_pagados = df_gastos[(df_gastos["mes"] == mes) & (df_gastos["año"] == año) & (df_gastos["estado"].str.lower() == "pagado")]
        deudas_mes = df_deudas[(df_deudas["mes"] == mes) & (df_deudas["año"] == año)]
        provisiones_mes = df_provisiones[(df_provisiones["mes"] == mes) & (df_provisiones["año"] == año)]
        ahorros_mes = df_ahorros[(df_ahorros["mes"] == mes) & (df_ahorros["año"] == año)]

        alerta_mostrada = False

        # 1. Provisiones "se usó" = Sí, pero monto usado = 0
        usadas_cero = provisiones_mes[(provisiones_mes["se_uso"].str.lower() == "si") & (provisiones_mes["monto_usado"] == 0)]
        if not usadas_cero.empty:
            st.error("⚠️ Hay provisiones marcadas como 'Se usó = Sí' pero sin monto registrado.")
            alerta_mostrada = True

        # 2. Ingresos < Gasto total
        ingreso_total = ing_mes["monto"].sum()
        gasto_normal = gastos_pagados["monto"].sum() + (deudas_mes["monto_cuota"] * deudas_mes["cuotas_mes"]).sum()
        gasto_provisiones = provisiones_mes["monto_usado"].sum()
        gasto_ahorros = ahorros_mes["monto_retirado"].sum()
        gasto_total = gasto_normal + gasto_provisiones + gasto_ahorros

        if ingreso_total < gasto_total:
            st.error("🚨 Gastaste más de lo que ganaste este mes.")
            alerta_mostrada = True

        # 3. Deudas con cuotas_mes = 0
        deudas_no_pagadas = deudas_mes[deudas_mes["cuotas_mes"] == 0]
        if not deudas_no_pagadas.empty:
            st.warning("🔔 Hay deudas sin cuotas registradas este mes.")
            alerta_mostrada = True

        # 4. Provisiones sin saldo (monto = 0)
        provisiones_sin_fondo = provisiones_mes[provisiones_mes["monto"] == 0]
        if not provisiones_sin_fondo.empty:
            st.warning("💡 Hay provisiones con saldo cero. Podrías no tener cómo cubrir futuros gastos.")
            alerta_mostrada = True

        if not alerta_mostrada:
            st.success("✨ Todo en orden este mes. ¡Buen trabajo!")

    except Exception as e:
        st.warning("No se pudieron evaluar las alertas.")
        st.text(f"Error: {e}")


with main_tabs[2]:
    # === Tabs principales reorganizados ===
    sub_tabs = st.tabs([
        "📥 Ingresos", 
        "🧾 Provisiones", 
        "💡 Gastos Fijos", 
        "🏦 Ahorros", 
        "👨‍👩‍👧‍👦 Reservas Familiares", 
        "💳 Deudas",
        "⚙️ Configuración"
    ])

    with sub_tabs[0]: mostrar_editor("Ingresos", columnas_dropdown=["cuenta"])
    with sub_tabs[1]: mostrar_editor("Provisiones")
    with sub_tabs[2]: mostrar_editor("Gastos Fijos", columnas_dropdown=["cuenta_pago"])
    with sub_tabs[3]: mostrar_editor("Ahorros", columnas_dropdown=["cuenta"])
    with sub_tabs[4]: mostrar_editor("Reservas Familiares", columnas_dropdown=["cuenta"])
    with sub_tabs[5]: mostrar_editor("Deudas")

    with sub_tabs[6]:
        st.subheader("🏦 Cuentas")
        try:
            df_cuentas = read_sheet_as_df(sheet, "Cuentas")
        except:
            st.warning("No se pudo cargar la hoja 'Cuentas'")
            df_cuentas = pd.DataFrame(columns=["nombre_cuenta", "banco", "tipo"])

        edited_cuentas = st.data_editor(
            df_cuentas,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True
        )

        if st.button("💾 Guardar cambios en Cuentas"):
            if edited_cuentas["nombre_cuenta"].isnull().any() or edited_cuentas["nombre_cuenta"].duplicated().any():
                st.error("No se permiten nombres vacíos ni duplicados.")
            else:
                write_df_to_sheet(sheet, "Cuentas", edited_cuentas)
                st.success("Cuentas actualizadas correctamente.")

    
with main_tabs[3]:
    st.subheader("📈 Reportes y Análisis")   
    
    rep_tabs = st.tabs([
        "💰 Ingresos vs Gastos",
        "📊 Distribución por Categoría",
        "📆 Evolución Mensual",
        "📤 Exportar Resumen"
    ])

    with rep_tabs[0]:
        st.markdown("💰 Ingresos vs Gastos Mensuales")

        try:
            df_ing = read_sheet_as_df(sheet, "Ingresos")
            df_gas = read_sheet_as_df(sheet, "Gastos Fijos")
            df_deu = read_sheet_as_df(sheet, "Deudas")
            df_pro = read_sheet_as_df(sheet, "Provisiones")
            df_aho = read_sheet_as_df(sheet, "Ahorros")

            # Agrupar ingresos
            df_ingresos = df_ing.groupby(["año", "mes"])["monto"].sum().reset_index(name="ingresos")

            # Gastos fijos pagados
            df_gas_pag = df_gas[df_gas["estado"].str.lower() == "pagado"]
            df_gastos_fijos = df_gas_pag.groupby(["año", "mes"])["monto"].sum().reset_index(name="gastos_fijos")

            # Deudas = monto_cuota * cuotas_mes
            df_deu["gasto_deuda"] = df_deu["monto_cuota"] * df_deu["cuotas_mes"]
            df_gastos_deuda = df_deu.groupby(["año", "mes"])["gasto_deuda"].sum().reset_index(name="deudas")

            # Provisiones usadas
            df_gastos_prov = df_pro.groupby(["año", "mes"])["monto_usado"].sum().reset_index(name="provisiones_usadas")

            # Ahorros retirados
            df_gastos_aho = df_aho.groupby(["año", "mes"])["monto_retirado"].sum().reset_index(name="ahorros_usados")

            # Combinar todo
            df_merge = df_ingresos \
                .merge(df_gastos_fijos, on=["año", "mes"], how="left") \
                .merge(df_gastos_deuda, on=["año", "mes"], how="left") \
                .merge(df_gastos_prov, on=["año", "mes"], how="left") \
                .merge(df_gastos_aho, on=["año", "mes"], how="left")

            # Reemplazar nulos por 0
            df_merge.fillna(0, inplace=True)

            df_merge["gastos_totales"] = df_merge["gastos_fijos"] + df_merge["deudas"] + df_merge["provisiones_usadas"] + df_merge["ahorros_usados"]
            df_merge["periodo"] = df_merge["mes"].astype(str).str.zfill(2) + "/" + df_merge["año"].astype(str)

            # === Gráfico en modo oscuro ===
            import matplotlib.pyplot as plt

            plt.style.use("dark_background")
            fig, ax = plt.subplots()
            ax.bar(df_merge["periodo"], df_merge["ingresos"], label="Ingresos", color="#4CAF50")
            ax.bar(df_merge["periodo"], df_merge["gastos_totales"], label="Gastos", color="#F44336", alpha=0.7)
            ax.set_title("Ingresos vs Gastos Totales por Mes")
            ax.set_ylabel("CLP")
            ax.legend()
            ax.tick_params(axis='x', rotation=45)

            st.pyplot(fig)

        except Exception as e:
            st.error("No se pudo generar el gráfico de ingresos vs gastos.")
            st.text(f"Error: {e}")

    with rep_tabs[1]:
        st.markdown("📊 Distribución del Gasto por Origen (Mes Actual)")
        try:
            # Volvemos a leer hojas por si se entra directamente
            df_gas = read_sheet_as_df(sheet, "Gastos Fijos")
            df_deu = read_sheet_as_df(sheet, "Deudas")
            df_pro = read_sheet_as_df(sheet, "Provisiones")
            df_aho = read_sheet_as_df(sheet, "Ahorros")

            # Filtro por mes/año
            gastos_pagados = df_gas[(df_gas["mes"] == mes) & (df_gas["año"] == año) & (df_gas["estado"].str.lower() == "pagado")]
            deudas_mes = df_deu[(df_deu["mes"] == mes) & (df_deu["año"] == año)]
            provisiones_mes = df_pro[(df_pro["mes"] == mes) & (df_pro["año"] == año)]
            ahorros_mes = df_aho[(df_aho["mes"] == mes) & (df_aho["año"] == año)]

            # Cálculos
            gasto_normal = gastos_pagados["monto"].sum() + (deudas_mes["monto_cuota"] * deudas_mes["cuotas_mes"]).sum()
            gasto_provisiones = provisiones_mes["monto_usado"].sum()
            gasto_ahorros = ahorros_mes["monto_retirado"].sum()

            # Preparar gráfico
            labels = ["Desde Ingresos Normales", "Desde Provisiones", "Desde Ahorros"]
            valores = [gasto_normal, gasto_provisiones, gasto_ahorros]

            # Si no hay gasto, no mostrar gráfico
            if sum(valores) == 0:
                st.info("No se han registrado gastos este mes.")
            else:
                import matplotlib.pyplot as plt
                plt.style.use("dark_background")
                fig, ax = plt.subplots()
                ax.pie(valores, labels=labels, autopct="%1.1f%%", startangle=90)
                ax.set_title("Distribución del Gasto por Origen")
                st.pyplot(fig)

        except Exception as e:
            st.error("No se pudo generar el gráfico de distribución.")
            st.text(f"Error: {e}")


    with rep_tabs[2]:  
        st.markdown("📆 Evolución Mensual de Ingresos, Gastos y Saldo Real")

        try:
            df_ing = read_sheet_as_df(sheet, "Ingresos")
            df_gas = read_sheet_as_df(sheet, "Gastos Fijos")
            df_deu = read_sheet_as_df(sheet, "Deudas")
            df_pro = read_sheet_as_df(sheet, "Provisiones")
            df_aho = read_sheet_as_df(sheet, "Ahorros")

            # Agrupar
            df_ingresos = df_ing.groupby(["año", "mes"])["monto"].sum().reset_index(name="ingresos")
            df_gas_pag = df_gas[df_gas["estado"].str.lower() == "pagado"]
            df_gastos_fijos = df_gas_pag.groupby(["año", "mes"])["monto"].sum().reset_index(name="gastos_fijos")

            df_deu["gasto_deuda"] = df_deu["monto_cuota"] * df_deu["cuotas_mes"]
            df_gastos_deuda = df_deu.groupby(["año", "mes"])["gasto_deuda"].sum().reset_index(name="deudas")

            df_gastos_prov = df_pro.groupby(["año", "mes"])["monto_usado"].sum().reset_index(name="provisiones_usadas")
            df_ahorros_ret = df_aho.groupby(["año", "mes"])["monto_retirado"].sum().reset_index(name="ahorros_usados")
            df_ahorros_ing = df_aho.groupby(["año", "mes"])["monto_ingreso"].sum().reset_index(name="ahorros_guardados")
            df_prov_guardadas = df_pro.groupby(["año", "mes"])["monto"].sum().reset_index(name="provisiones_guardadas")

            # Merge general
            df_merge = df_ingresos \
                .merge(df_gastos_fijos, on=["año", "mes"], how="left") \
                .merge(df_gastos_deuda, on=["año", "mes"], how="left") \
                .merge(df_gastos_prov, on=["año", "mes"], how="left") \
                .merge(df_ahorros_ret, on=["año", "mes"], how="left") \
                .merge(df_ahorros_ing, on=["año", "mes"], how="left") \
                .merge(df_prov_guardadas, on=["año", "mes"], how="left")

            df_merge.fillna(0, inplace=True)

            # Cálculos
            df_merge["gastos_totales"] = df_merge["gastos_fijos"] + df_merge["deudas"] + df_merge["provisiones_usadas"] + df_merge["ahorros_usados"]
            df_merge["saldo_real"] = df_merge["ingresos"] - df_merge["gastos_totales"] - df_merge["provisiones_guardadas"] - df_merge["ahorros_guardados"]
            df_merge["periodo"] = df_merge["mes"].astype(str).str.zfill(2) + "/" + df_merge["año"].astype(str)

            # Gráfico
            import matplotlib.pyplot as plt
            plt.style.use("dark_background")
            fig, ax = plt.subplots()
            ax.plot(df_merge["periodo"], df_merge["ingresos"], label="Ingresos", marker="o", color="#4CAF50")
            ax.plot(df_merge["periodo"], df_merge["gastos_totales"], label="Gastos Totales", marker="o", color="#F44336")
            ax.plot(df_merge["periodo"], df_merge["saldo_real"], label="Saldo Disponible Real", marker="o", color="#2196F3")

            ax.set_title("Evolución de Ingresos, Gastos y Saldo Real")
            ax.set_ylabel("CLP")
            ax.set_xlabel("Mes/Año")
            ax.tick_params(axis="x", rotation=45)
            ax.legend()
            st.pyplot(fig)

        except Exception as e:
            st.error("No se pudo generar el gráfico de evolución.")
            st.text(f"Error: {e}")


    with rep_tabs[3]:  
        st.markdown(" 📤 Exportar tus datos")

        try:
            hojas = ["Ingresos", "Gastos Fijos", "Deudas", "Provisiones", "Ahorros", "Reservas Familiares"]
            dfs_mes = {}
            dfs_año = {}

            for hoja in hojas:
                df = df_hojas.get(hoja, pd.DataFrame())

                if "mes" in df.columns and "año" in df.columns:
                    df_mes = df[(df["mes"] == mes) & (df["año"] == año)]
                    df_año = df[df["año"] == año]
                else:
                    df_mes = df.copy()
                    df_año = df.copy()

                dfs_mes[hoja] = df_mes
                dfs_año[hoja] = df_año

            # === Botón para descargar resumen mensual
            if st.button("📥 Descargar resumen mensual en Excel"):
                output = io.BytesIO()
                wb = Workbook()
                for nombre, df in dfs_mes.items():
                    ws = wb.create_sheet(title=nombre[:31])
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                wb.remove(wb["Sheet"])
                wb.save(output)
                st.download_button(
                    label="⬇️ Descargar archivo mensual",
                    data=output.getvalue(),
                    file_name=f"Resumen_{mes}_{año}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            # === Botón para descargar resumen anual
            if st.button("📥 Descargar histórico anual en Excel"):
                output = io.BytesIO()
                wb = Workbook()
                for nombre, df in dfs_año.items():
                    ws = wb.create_sheet(title=nombre[:31])
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                wb.remove(wb["Sheet"])
                wb.save(output)
                st.download_button(
                    label="⬇️ Descargar archivo anual",
                    data=output.getvalue(),
                    file_name=f"Resumen_Anual_{año}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error("No se pudo generar el archivo para exportar.")
            st.text(f"Error: {e}")
        
with main_tabs[4]:
    st.subheader("🧮 Simulador de Próximo Mes")

    # Calcular próximo mes
    nuevo_mes, nuevo_año = obtener_mes_siguiente(mes, año)
    st.markdown(f" Simulación para: **{nuevo_mes}/{nuevo_año}**")

    # === Valores base desde el mes actual
    df_ing = df_hojas["Ingresos"]
    df_gas = df_hojas["Gastos Fijos"]
    df_deu = df_hojas["Deudas"]
    df_pro = df_hojas["Provisiones"]
    df_aho = df_hojas["Ahorros"]

    ingreso_real = df_ing[(df_ing["mes"] == mes) & (df_ing["año"] == año)]["monto"].sum()
    gastos_fijos_reales = df_gas[(df_gas["mes"] == mes) & (df_gas["año"] == año) & (df_gas["estado"].str.lower() == "pagado")]["monto"].sum()
    provisiones_real = df_pro[(df_pro["mes"] == mes) & (df_pro["año"] == año)]["monto"].sum()
    ahorro_real = df_aho[(df_aho["mes"] == mes) & (df_aho["año"] == año)]["monto_ingreso"].sum()
    deudas_real = (df_deu[(df_deu["mes"] == mes) & (df_deu["año"] == año)]["monto_cuota"] * df_deu["cuotas_mes"]).sum()

    # Entradas del usuario
    ingreso_simulado = st.number_input("💰 Ingreso estimado", min_value=0, value=int(ingreso_real), step=10000)
    gasto_estimado = st.number_input("💸 Gastos fijos esperados", min_value=0, value=int(gastos_fijos_reales), step=10000)
    provisiones_estimadas = st.number_input("🏷️ Provisiones a guardar", min_value=0, value=int(provisiones_real), step=10000)
    ahorro_estimado = st.number_input("🏦 Ahorro previsto", min_value=0, value=int(ahorro_real), step=10000)
    deuda_estimadas = st.number_input("💳 Pago de deudas estimado", min_value=0, value=int(deudas_real), step=10000)

    # Cálculos
    gasto_total = gasto_estimado + deuda_estimadas
    saldo_proyectado = ingreso_simulado - gasto_total - provisiones_estimadas - ahorro_estimado

    # Métricas
    st.markdown(" 📊 Resultado Proyectado")
    col1, col2 = st.columns(2)
    col1.metric("💸 Gasto Total", f"${gasto_total:,.0f}")
    col2.metric("🧮 Saldo Disponible", f"${saldo_proyectado:,.0f}")

    # === Gráfico de distribución proyectada del ingreso ===
    st.markdown(" 🧁 Distribución del ingreso simulado")

    restante = max(ingreso_simulado - gasto_total - provisiones_estimadas - ahorro_estimado, 0)

    etiquetas = ["Gastos fijos + Deudas", "Provisiones", "Ahorros", "Saldo Libre"]
    valores = [gasto_total, provisiones_estimadas, ahorro_estimado, restante]

    if ingreso_simulado == 0 or sum(valores) == 0:
        st.info("No hay ingreso simulado para mostrar la distribución.")
    else:
        import matplotlib.pyplot as plt
        plt.style.use("dark_background")
        fig, ax = plt.subplots()
        ax.pie(valores, labels=etiquetas, autopct="%1.1f%%", startangle=90)
        ax.set_title("Distribución proyectada del ingreso")
        st.pyplot(fig)


